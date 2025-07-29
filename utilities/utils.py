from openpyxl.reader.excel import load_workbook


class utlis:
    def read_data_from_excel(filename,sheet):
        datalist = []
        workbook = load_workbook(filename=filename)
        sheet = workbook[sheet]
        row_count = sheet.max_row
        column_count = sheet.max_column

        for i in range(2, row_count+1):
            row = []
            for j in range(1,column_count+1):
                row.append(sheet.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist




